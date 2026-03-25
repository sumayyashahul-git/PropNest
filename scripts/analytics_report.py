import requests
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AUTH_URL = "http://localhost:5002/api"
PROPERTY_URL = "http://localhost:5001/api"
ANALYTICS_URL = "http://localhost:5003/api"

PROPERTY_TYPES = ["Apartment", "Villa", "Townhouse", "Penthouse", "Studio", "Office"]
LISTING_TYPES = ["For Sale", "For Rent"]
EMIRATES = ["Dubai", "Abu Dhabi", "Sharjah", "Ajman", "Other"]

def get_token():
    response = requests.post(f"{AUTH_URL}/Auth/login", json={
        "email": "admin@propnest.ae",
        "password": "Admin@123"
    })
    if response.status_code == 200:
        return response.json()["token"]
    return None

def fetch_all_properties(token):
    headers = {"Authorization": f"Bearer {token}"}
    all_props = []
    page = 1

    while True:
        response = requests.get(
            f"{PROPERTY_URL}/Properties",
            headers=headers,
            params={"page": page, "pageSize": 50}
        )
        if response.status_code != 200:
            break

        data = response.json()
        items = data.get("items", [])
        all_props.extend(items)

        if len(all_props) >= data.get("total", 0):
            break
        page += 1

    return all_props

def fetch_analytics(token):
    headers = {"Authorization": f"Bearer {token}"}

    stats = requests.get(
        f"{ANALYTICS_URL}/Analytics/dashboard",
        headers=headers
    ).json()

    top_viewed = requests.get(
        f"{ANALYTICS_URL}/Analytics/top-viewed",
        headers=headers
    ).json()

    recent = requests.get(
        f"{ANALYTICS_URL}/Analytics/recent-listings",
        headers=headers
    ).json()

    return stats, top_viewed, recent

def style_header_row(ws, row, bg_color):
    fill = PatternFill(start_color=bg_color,
                       end_color=bg_color,
                       fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

def auto_fit_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

def generate_report():
    print("\n📊 PropNest Analytics Report Generator")
    print("="*50)

    print("\n🔐 Authenticating...")
    token = get_token()
    if not token:
        print("❌ Failed to authenticate!")
        return

    print("✅ Authenticated\n")
    print("📡 Fetching data from APIs...")

    properties = fetch_all_properties(token)
    stats, top_viewed, recent_listings = fetch_analytics(token)

    print(f"✅ Fetched {len(properties)} properties")

    df = pd.DataFrame(properties)

    if df.empty:
        print("❌ No properties found!")
        return

    df["typeName"] = df["type"].apply(
        lambda x: PROPERTY_TYPES[x] if x < len(PROPERTY_TYPES) else "Unknown")
    df["listingTypeName"] = df["listingType"].apply(
        lambda x: LISTING_TYPES[x] if x < len(LISTING_TYPES) else "Unknown")
    df["createdAt"] = pd.to_datetime(df["createdAt"]).dt.strftime("%Y-%m-%d")

    filename = f"PropNest_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:

        # ── Sheet 1: Summary ──────────────────────────
        summary_data = {
            "Metric": [
                "Total Properties",
                "Total Views",
                "Price Drops",
                "Average Price (AED)",
                "Highest Price (AED)",
                "Lowest Price (AED)",
                "For Sale",
                "For Rent",
                "Report Generated"
            ],
            "Value": [
                stats.get("totalProperties", len(properties)),
                stats.get("totalViews", 0),
                stats.get("totalPriceDrops", 0),
                f"AED {df['price'].mean():,.0f}",
                f"AED {df['price'].max():,.0f}",
                f"AED {df['price'].min():,.0f}",
                len(df[df["listingType"] == 0]),
                len(df[df["listingType"] == 1]),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]
        }
        pd.DataFrame(summary_data).to_excel(
            writer, sheet_name="Summary", index=False)
        style_header_row(writer.sheets["Summary"], 1, "1E3A5F")
        auto_fit_columns(writer.sheets["Summary"])

        # ── Sheet 2: All Properties ───────────────────
        columns = ["title", "typeName", "listingTypeName",
                   "location", "emirate", "price",
                   "bedrooms", "bathrooms", "areaSqFt",
                   "viewCount", "agentName", "createdAt"]

        df[columns].rename(columns={
            "title": "Title",
            "typeName": "Type",
            "listingTypeName": "Listing",
            "location": "Location",
            "emirate": "Emirate",
            "price": "Price (AED)",
            "bedrooms": "Beds",
            "bathrooms": "Baths",
            "areaSqFt": "Area (sqft)",
            "viewCount": "Views",
            "agentName": "Agent",
            "createdAt": "Listed Date"
        }).to_excel(writer, sheet_name="All Properties", index=False)
        style_header_row(writer.sheets["All Properties"], 1, "C0392B")
        auto_fit_columns(writer.sheets["All Properties"])

        # ── Sheet 3: By Type ──────────────────────────
        by_type = df.groupby("typeName").agg(
            Count=("id", "count"),
            Avg_Price=("price", "mean"),
            Min_Price=("price", "min"),
            Max_Price=("price", "max"),
            Total_Views=("viewCount", "sum")
        ).round(0).reset_index()
        by_type.columns = ["Type", "Count",
                           "Avg Price (AED)", "Min Price (AED)",
                           "Max Price (AED)", "Total Views"]
        by_type.to_excel(writer, sheet_name="By Type", index=False)
        style_header_row(writer.sheets["By Type"], 1, "27AE60")
        auto_fit_columns(writer.sheets["By Type"])

        # ── Sheet 4: By Emirate ───────────────────────
        by_emirate = df.groupby("emirate").agg(
            Count=("id", "count"),
            Avg_Price=("price", "mean"),
            Total_Views=("viewCount", "sum")
        ).round(0).reset_index()
        by_emirate.columns = ["Emirate", "Count",
                              "Avg Price (AED)", "Total Views"]
        by_emirate.to_excel(writer, sheet_name="By Emirate", index=False)
        style_header_row(writer.sheets["By Emirate"], 1, "8E44AD")
        auto_fit_columns(writer.sheets["By Emirate"])

        # ── Sheet 5: Top Viewed ───────────────────────
        if top_viewed:
            df_top = pd.DataFrame(top_viewed)
            df_top[["title", "location", "currentPrice",
                    "viewCount", "listedAt"]].rename(columns={
                "title": "Title",
                "location": "Location",
                "currentPrice": "Price (AED)",
                "viewCount": "Views",
                "listedAt": "Listed At"
            }).to_excel(writer, sheet_name="Top Viewed", index=False)
            style_header_row(writer.sheets["Top Viewed"], 1, "E67E22")
            auto_fit_columns(writer.sheets["Top Viewed"])

    print(f"\n✅ Report saved: {filename}")
    print(f"\n📋 Sheets created:")
    print(f"   1. Summary          → key metrics")
    print(f"   2. All Properties   → {len(df)} listings")
    print(f"   3. By Type          → breakdown by property type")
    print(f"   4. By Emirate       → breakdown by emirate")
    print(f"   5. Top Viewed       → most popular listings")
    print(f"\n{'='*50}\n")

if __name__ == "__main__":
    generate_report()
